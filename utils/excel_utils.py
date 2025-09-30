import openpyxl




def read_excel():
    # 打开 excel 文件
    workbook = openpyxl.load_workbook("./data/api_testcases.xlsx")  # 这里替换成你的文件路径
    # workbook = openpyxl.load_workbook("../data/api_testcases.xlsx")

    # 选择表
    worksheet = workbook["Sheet1"]

    # 读取数据操作
    # 思路：因为 dict(zip(keys, value)) 可以把读取到的数据变成字典类型
    data = []  # 空列表，用于组装字典
    keys = [cell.value for cell in worksheet[1]]  # 获取 key 行，也就是表的第一行，生成一个key列表

    for row in worksheet.iter_rows(min_row=2, values_only=True): # 从第二行开始逐行读取，只返回值
        dict_data = dict(zip(keys, row))  # 组装成字典
        # data.append(dict_data)
        #如果读取的is_true是TRUE则append，否则，不append
        # print(dict_data["is_ture"])
        if dict_data["is_true"]:
            data.append(dict_data)


    # 打印拿到的所有数据
    print(data)

    # 关闭 excel 文件
    workbook.close()

    return data

# read_excel()